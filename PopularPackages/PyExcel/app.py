import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
column = sheet["a"]  # this will return all the stuff in the a column

# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

# this code is useful for iterate through columns
# sheet.cell(row=1, column=1) same as line 9

# print(sheet.max_row)
# print(sheet.max_column)

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

print(column)
cells = sheet["a:c"]
print(cells)
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])  # adds row to the end of the sheet
wb.save("transactions2.xlsx")
